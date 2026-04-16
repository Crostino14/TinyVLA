"""
analyze_embedding_overlap_rollout.py
=====================================

Overview
--------
This script measures how semantically similar TinyVLA's internal representations
are when the same robot task is described with different natural language commands
(linguistic variation levels L1, L2, L3 vs. the default full command).

The core idea is:
  If the VLA policy truly understands language, a paraphrase of a command should
  produce an internal embedding close to the original. If the embeddings diverge
  despite similar meaning, the model may be performing shallow pattern matching
  rather than genuine language comprehension.

Two complementary similarity metrics are computed for each default/variation pair:

  1. **Semantic similarity** — measured on the model's own hidden state vectors:
       - Cosine Similarity   : measures the angle between embedding vectors
                               (1.0 = identical direction, 0.0 = orthogonal)
       - Euclidean Distance  : measures the L2 distance in embedding space
                               (0.0 = identical, larger = more different)

  2. **Lexical similarity** — measured on the command strings themselves:
       - Normalized Levenshtein Distance : character-level edit distance,
         normalized to [0, 1] (0.0 = identical strings, 1.0 = completely different)

  Comparing semantic and lexical similarity together reveals whether the model's
  representations track meaning (semantic) independently of surface form (lexical).

Embedding Source
----------------
The embeddings analyzed here are the **hidden_states** (last-layer activations)
of the GPT-NeoX/Pythia backbone inside TinyVLA's LLaVA-Pythia architecture,
captured at the point BEFORE the diffusion action head. Specifically:
  - Extracted during actual rollout execution (not in a text-only forward pass),
    so they encode the full multimodal context: image + language + robot state.
  - Mean-pooled over the token sequence dimension to produce a single fixed-size
    vector per inference step.
  - Two embedding granularities are stored:
      * `embedding`             : Mean over ALL steps and ALL rollouts (1 vector)
      * `embedding_per_rollout` : Mean over steps within each rollout (N vectors)

Supported Modes
---------------
  - **first_step_only**: Each rollout contributes one embedding from its first
    observation. Useful for analyzing the language encoding at episode start,
    before any action is taken.
  - **full rollout**   : Each rollout contributes embeddings averaged over all
    timesteps. Provides a more representative encoding of the full episode's
    language-conditioned behavior.

Linguistic Variation Levels
----------------------------
  DEFAULT : Original task command used during training
            Example: "Put the bowl on the stove"
  L1      : Lexical paraphrase (synonym substitution)
            Example: "Set the bowl on the stove"
  L2      : Structural/passive reformulation
            Example: "The stove needs to have the bowl on it"
  L3      : Indirect referential expression
            Example: "Put the object between the wine bottle and the cream
                      cheese on the stove"

Input File Format
-----------------
Embedding data is stored as Python pickle files (.pkl). Each file is a
dictionary mapping string keys to per-task embedding records:

  Key format   : "task_{id:02d}_{level}"
                 e.g., "task_00_default", "task_03_l2"

  Value format : dict with the following fields:
    - "embedding"            (np.ndarray) : Mean embedding over all rollouts+steps
    - "embedding_per_rollout"(np.ndarray) : Shape (N_rollouts, embed_dim)
    - "command_text"         (str)        : The language command used
    - "command_level"        (str)        : "default", "l1", "l2", or "l3"
    - "task_id"              (int)        : 0-based task index
    - "model"                (str)        : Model identifier
    - "num_rollouts"         (int)        : Number of rollouts collected
    - "first_step_only"      (bool)       : Whether first-step mode was used
    - "total_steps"          (int)        : Total inference steps recorded
    - "success_rate"         (float)      : Episode success rate (0.0–1.0)

Output Files
------------
For each analysis run, three types of files are generated:
  1. `<output_base>.csv`                  — Raw results table (all tasks & levels)
  2. `<output_base>_overlap_l1.xlsx`      — Formatted L1 variation overlap table
  3. `<output_base>_overlap_l2.xlsx`      — Formatted L2 variation overlap table
  4. `<output_base>_overlap_l3.xlsx`      — Formatted L3 variation overlap table

Dependencies
------------
  - numpy           : Array operations and statistics
  - sklearn         : cosine_similarity and euclidean_distances computation
  - pandas          : Results aggregation and CSV export
  - Levenshtein     : Fast C-based Levenshtein distance computation
  - openpyxl        : Formatted Excel table generation
  - pickle          : Deserialization of embedding files
  - glob            : Auto-discovery of embedding files in a directory
  - argparse        : Command-line argument parsing

Usage
-----
Single file:
    python analyze_embedding_overlap_rollout.py --embedding_file embeddings.pkl

Multiple files (one per level, recommended):
    python analyze_embedding_overlap_rollout.py --embedding_files \\
        .../rollout_embeddings_tinyvla_libero_goal_default_first_step_r10.pkl \\
        .../rollout_embeddings_tinyvla_libero_goal_l1_first_step_r10.pkl \\
        .../rollout_embeddings_tinyvla_libero_goal_l2_first_step_r10.pkl \\
        .../rollout_embeddings_tinyvla_libero_goal_l3_first_step_r10.pkl

Directory mode:
    python analyze_embedding_overlap_rollout.py \\
        --embedding_dir /mnt/beegfs/a.cardamone7/outputs/embeddings/tinyvla

Per-rollout analysis for a specific task:
    python analyze_embedding_overlap_rollout.py --embedding_files ... \\
        --per_rollout_analysis --task_id 3 --level l2
"""

import os       # Filesystem path construction and directory operations
import glob     # Unix-style wildcard file pattern matching for auto-discovery
import pickle   # Serialization/deserialization of Python objects (embedding dicts)
import numpy as np  # Numerical arrays for embedding vectors and statistics

from sklearn.metrics.pairwise import (
    cosine_similarity,   # Computes cosine similarity between pairs of vectors
    euclidean_distances, # Computes L2 (Euclidean) distance between pairs of vectors
)
import pandas as pd  # Tabular data aggregation, statistics, and CSV export

from Levenshtein import distance as levenshtein_distance
# Fast C-extension implementation of the Levenshtein (edit) distance algorithm.
# Computes the minimum number of single-character insertions, deletions, and
# substitutions required to transform one string into another.


# =============================================================================
# Metric computation utilities
# =============================================================================

def compute_levenshtein_normalized(text1, text2):
    """
    Compute the normalized Levenshtein (edit) distance between two strings.

    The raw Levenshtein distance counts the minimum number of character-level
    edits (insertions, deletions, substitutions) needed to transform `text1`
    into `text2`. This function normalizes that count by the length of the
    longer string, producing a value in [0, 1] that is comparable across
    command pairs of different lengths.

    Normalization rationale:
      - A raw distance of 10 is very significant for a 12-character string
        but negligible for a 200-character string. Dividing by the maximum
        length makes the metric length-invariant.
      - Dividing by `max(len(text1), len(text2))` gives the fraction of the
        longer string that would need to be rewritten to produce the other.

    Parameters
    ----------
    text1 : str
        The first command string (typically the default task command).
    text2 : str
        The second command string (a linguistic variation: L1, L2, or L3).

    Returns
    -------
    float
        Normalized Levenshtein distance in the range [0.0, 1.0]:
          - 0.0 : The two strings are identical (zero edits needed)
          - 1.0 : The strings are completely different (all characters differ)
        Returns 0.0 if both strings are empty (edge case guard).

    Examples
    --------
    >>> compute_levenshtein_normalized("Put the bowl on the stove",
    ...                                "Set the bowl on the stove")
    0.04  # Only "Put" → "Set": 2 substitutions out of 25 characters

    >>> compute_levenshtein_normalized("Turn on the stove",
    ...                                "The stove needs to be turned on")
    0.58  # Major structural change: high edit distance
    """
    # Compute the raw integer edit distance between the two input strings
    lev_dist = levenshtein_distance(text1, text2)

    # Normalize by the length of the longer string to get a [0, 1] value
    max_len = max(len(text1), len(text2))

    # Guard against division by zero when both strings are empty
    if max_len == 0:
        return 0.0

    return lev_dist / max_len


# =============================================================================
# Data loading utilities
# =============================================================================

def load_embeddings(embedding_files=None, embedding_dir=None):
    """
    Load and merge TinyVLA rollout embedding records from one or more pickle files.

    Each pickle file is expected to contain a dict mapping string keys
    (e.g., "task_03_l2") to embedding record dicts. This function loads all
    files and merges them into a single combined dictionary, supporting both
    a directory-based auto-discovery workflow and an explicit file list workflow.

    Use cases:
      - A single `.pkl` file containing all levels for all tasks.
      - One `.pkl` file per command level (recommended), each containing
        all tasks for that level. Files are merged by dict update.
      - Auto-discovery of all `.pkl` files in a directory.

    Parameters
    ----------
    embedding_files : list of str, optional
        Explicit list of absolute or relative pickle file paths to load.
        Used in CLI `--embedding_files` mode.
    embedding_dir : str, optional
        Path to a directory. All `.pkl` files found via `glob` are loaded
        and merged. Files are sorted alphabetically before loading.
        Takes priority over `embedding_files` if both are provided.

    Returns
    -------
    dict
        Merged dictionary of all embedding records from all files.
        Keys are strings like "task_00_default", "task_03_l2", etc.
        Values are per-task embedding dicts (see module docstring for schema).

    Raises
    ------
    ValueError
        If neither `embedding_files` nor `embedding_dir` is provided.

    Notes
    -----
    - If the same key appears in multiple files, the LAST file's value wins
      due to Python's `dict.update()` behavior. To avoid silent data loss,
      store each command level in a separate file with non-overlapping keys.
    - Files are loaded in sorted order (alphabetical) to ensure deterministic
      merge behavior across platforms.
    """
    all_embeddings = {}  # Accumulator: merged embedding dict across all files

    # ── Determine the list of files to load ───────────────────────────────
    if embedding_dir:
        # Auto-discover all pickle files in the directory, sorted alphabetically
        files = sorted(glob.glob(os.path.join(embedding_dir, "*.pkl")))
        print(f"Found {len(files)} pickle files in {embedding_dir}")
    elif embedding_files:
        # Use the explicitly provided file list
        files = embedding_files
    else:
        # Neither source was provided: raise a descriptive error
        raise ValueError("Must provide embedding_files or embedding_dir")

    # ── Load and merge each file ───────────────────────────────────────────
    for filepath in files:
        print(f"  Loading: {os.path.basename(filepath)}")  # Show only filename for brevity
        with open(filepath, 'rb') as f:
            data = pickle.load(f)  # Deserialize the pickle file into a Python dict
            all_embeddings.update(data)  # Merge: later files overwrite duplicate keys

    print(f"\nTotal embeddings loaded: {len(all_embeddings)}")
    return all_embeddings


# =============================================================================
# Main analysis function
# =============================================================================

def analyze_embeddings(embeddings, output_csv="analysis_results.csv"):
    """
    Compute and report semantic and lexical distances between default task commands
    and their L1/L2/L3 linguistic variation counterparts.

    For each LIBERO task and each variation level (L1, L2, L3), this function:
      1. Retrieves the mean rollout embedding for the default and variation commands.
      2. Computes three distance/similarity metrics between the embedding pair.
      3. Aggregates per-level and overall statistics (mean ± std).
      4. Saves all results to a CSV file.
      5. Exports formatted per-level overlap tables to `.xlsx` files.

    Metric interpretation guide:
    ──────────────────────────────
    - **Cosine Similarity** (0.0–1.0):
        Measures the cosine of the angle between two embedding vectors.
        High values (→ 1.0) mean the model encodes both commands similarly.
        Low values indicate the model produces distinct internal representations.
        This is the primary semantic metric.

    - **Euclidean Distance** (0.0–∞):
        Measures the L2 norm between two embedding vectors.
        Complements cosine similarity by capturing magnitude differences,
        not just directional alignment.

    - **Levenshtein Distance** (0.0–1.0):
        Measures character-level string edit distance, normalized by max length.
        High values indicate the strings are lexically very different.
        Comparing this to cosine similarity reveals whether semantic similarity
        tracks or diverges from surface form similarity.

    Parameters
    ----------
    embeddings : dict
        Dictionary of embedding records (output of `load_embeddings()`).
        Keys: "task_{id:02d}_{level}" strings.
        Values: dicts with "embedding", "command_text", "command_level",
        "task_id", "num_rollouts", "success_rate", "first_step_only", etc.
    output_csv : str, optional
        Output path for the raw results CSV file.
        Default: "analysis_results.csv"

    Returns
    -------
    pandas.DataFrame
        Results DataFrame with one row per (task_id, level) pair and columns:
          - task_id, level
          - default_command, variation_command
          - default_rollouts, variation_rollouts
          - default_success_rate, variation_success_rate
          - cosine_similarity    (float, 0.0–1.0)
          - euclidean_distance   (float, 0.0–∞)
          - levenshtein_distance (float, 0.0–1.0)

    Notes
    -----
    Embedding Flattening:
      Both the default and variation embeddings are flattened to 1D before
      computing pairwise distances. `sklearn`'s `cosine_similarity` and
      `euclidean_distances` expect 2D inputs, so each 1D vector is wrapped
      in a list: `cosine_similarity([vec1], [vec2])[0, 0]`.

    NaN Safety:
      If a task's `default` embedding is missing from the data, the task is
      silently skipped (the `if 'default' not in task_data: continue` guard).
      Missing variation levels are reported as "[NOT FOUND]" in the console.
    """
    print(f"\nAnalyzing {len(embeddings)} rollout-based mean embeddings\n")

    # ── Diagnostic info: print metadata from the first embedding record ────
    if embeddings:
        first_key = next(iter(embeddings.keys()))   # Get any key from the dict
        first_data = embeddings[first_key]           # Get its embedding record

        print(f"  Model: {first_data.get('model', 'unknown')}")
        print(f"  Mean embedding shape: {first_data['embedding'].shape}")
        # embedding_per_rollout has shape (N_rollouts, embed_dim)
        print(f"  Embeddings per rollout shape: {first_data['embedding_per_rollout'].shape}")
        print(f"  Number of rollouts: {first_data.get('num_rollouts', 'N/A')}")

        # Determine and display the extraction mode
        mode = "First step only" if first_data.get('first_step_only', False) else "Full rollout"
        print(f"  Mode: {mode}")

        if 'total_steps' in first_data:
            print(f"  Total steps: {first_data['total_steps']}")
        if 'success_rate' in first_data:
            print(f"  Success rate: {first_data['success_rate']:.2%}")  # Display as percentage
        print()

    # ── Organize embeddings by task_id ─────────────────────────────────────
    # tasks[task_id][command_level] → embedding record
    tasks = {}
    for key, data in embeddings.items():
        task_id = data['task_id']
        if task_id not in tasks:
            tasks[task_id] = {}  # Initialize inner dict for this task
        # Nest by command level ("default", "l1", "l2", "l3")
        tasks[task_id][data['command_level']] = data

    results = []  # Will accumulate one dict per (task, level) comparison

    print("=" * 80)
    print("DISTANCE ANALYSIS: Semantic (Cosine) + Lexical (Levenshtein)")
    print("Using ROLLOUT MEAN EMBEDDINGS (pre-action-head hidden_states)")
    print("=" * 80)

    # ── Process each task in ascending ID order ────────────────────────────
    for task_id in sorted(tasks.keys()):
        task_data = tasks[task_id]  # Dict of command_level → embedding record

        # Skip tasks that have no default embedding (can't compute relative distances)
        if 'default' not in task_data:
            continue

        # ── Retrieve default embedding and metadata ─────────────────────────
        default_emb = task_data['default']['embedding']
        # Ensure 1D: flatten multi-dimensional arrays (e.g., shape (1, D) → (D,))
        if default_emb.ndim > 1:
            default_emb = default_emb.flatten()

        default_cmd = task_data['default']['command_text']           # Original command string
        default_rollouts = task_data['default'].get('num_rollouts', 'N/A')  # Rollout count
        default_sr = task_data['default'].get('success_rate', None)  # Success rate or None

        print(f"\nTask {task_id}")
        # Format success rate as percentage if available, otherwise empty string
        sr_str = f", SR={default_sr:.0%}" if default_sr is not None else ""
        print(f"  Default: {default_cmd} [{default_rollouts} rollouts{sr_str}]")

        # ── Compare default against each variation level ────────────────────
        for level in ['l1', 'l2', 'l3']:
            if level not in task_data:
                # Variation data missing: log and skip (no CSV entry for missing data)
                print(f"  {level.upper():3s}:     [NOT FOUND]")
                continue

            # ── Retrieve variation embedding and metadata ───────────────────
            var_emb = task_data[level]['embedding']
            if var_emb.ndim > 1:
                var_emb = var_emb.flatten()  # Flatten to 1D for sklearn functions

            var_cmd = task_data[level]['command_text']
            var_rollouts = task_data[level].get('num_rollouts', 'N/A')
            var_sr = task_data[level].get('success_rate', None)

            # ── Metric 1: Cosine Similarity ─────────────────────────────────
            # sklearn expects 2D arrays: wrap each 1D vector in a list
            # Result is a (1, 1) matrix; [0, 0] extracts the scalar value
            cos_sim = cosine_similarity([default_emb], [var_emb])[0, 0]

            # ── Metric 2: Euclidean (L2) Distance ───────────────────────────
            # Same 2D wrapping as cosine_similarity
            euc_dist = euclidean_distances([default_emb], [var_emb])[0, 0]

            # ── Metric 3: Normalized Levenshtein Distance ───────────────────
            # Operates on raw command strings, not embeddings
            lev_dist = compute_levenshtein_normalized(default_cmd, var_cmd)

            # ── Accumulate result for this (task, level) pair ───────────────
            results.append({
                'task_id': task_id,
                'level': level,
                'default_command': default_cmd,
                'variation_command': var_cmd,
                'default_rollouts': default_rollouts,
                'variation_rollouts': var_rollouts,
                'default_success_rate': default_sr,
                'variation_success_rate': var_sr,
                'cosine_similarity': cos_sim,
                'euclidean_distance': euc_dist,
                'levenshtein_distance': lev_dist,
            })

            # ── Console output for this comparison ──────────────────────────
            sr_str = f", SR={var_sr:.0%}" if var_sr is not None else ""
            print(f"  {level.upper():3s}:     {var_cmd} [{var_rollouts} rollouts{sr_str}]")
            print(f"           Semantic:   Cosine_sim={cos_sim:.4f}")
            print(f"           Semantic:   Euclidean_dist(L2)={euc_dist:.4f}")
            print(f"           Lexical:    Lev_dist={lev_dist:.4f}")

    # ── Build results DataFrame ────────────────────────────────────────────
    df = pd.DataFrame(results)  # One row per (task_id, level) comparison

    # ── Per-level aggregate statistics ────────────────────────────────────
    print("\n" + "=" * 80)
    print("AVERAGE STATISTICS BY LEVEL")
    print("=" * 80)

    for level in ['l1', 'l2', 'l3']:
        level_data = df[df['level'] == level]  # Filter to this level's rows
        if len(level_data) == 0:
            continue  # Skip levels with no data (no rows in the DataFrame)

        print(f"\n{level.upper()}:")
        # .mean() and .std() use ddof=1 (sample std) by default in pandas
        print(
            f"  Cosine Similarity (semantic):     "
            f"{level_data['cosine_similarity'].mean():.4f} ± "
            f"{level_data['cosine_similarity'].std():.4f}"
        )
        print(
            f"  Euclidean Distance L2 (semantic): "
            f"{level_data['euclidean_distance'].mean():.4f} ± "
            f"{level_data['euclidean_distance'].std():.4f}"
        )
        print(
            f"  Levenshtein Distance (lexical):   "
            f"{level_data['levenshtein_distance'].mean():.4f} ± "
            f"{level_data['levenshtein_distance'].std():.4f}"
        )

    # ── Overall aggregate statistics (across all levels) ───────────────────
    print("\n" + "=" * 80)
    print("OVERALL STATISTICS")
    print("=" * 80)
    print(
        f"  Overall Cosine Similarity:      "
        f"{df['cosine_similarity'].mean():.4f} ± {df['cosine_similarity'].std():.4f}"
    )
    print(
        f"  Overall Euclidean Distance L2:  "
        f"{df['euclidean_distance'].mean():.4f} ± {df['euclidean_distance'].std():.4f}"
    )
    print(
        f"  Overall Levenshtein Distance:   "
        f"{df['levenshtein_distance'].mean():.4f} ± {df['levenshtein_distance'].std():.4f}"
    )

    # ── Save raw results to CSV ────────────────────────────────────────────
    df.to_csv(output_csv, index=False)  # index=False: don't write row numbers to CSV
    print(f"\n✓ Results saved: {output_csv}")

    # ── Save formatted Excel overlap tables (one per level) ────────────────
    # Strip the ".csv" extension to build a shared base path for Excel files
    base_path = output_csv.rsplit(".", 1)[0]

    for lvl in ['l1', 'l2', 'l3']:
        level_df = df[df['level'] == lvl].reset_index(drop=True)  # Filter + reindex
        if len(level_df) == 0:
            continue  # Skip levels with no data

        # Construct the Excel output path: e.g., "combined_analysis_overlap_l1.xlsx"
        formatted_path = f"{base_path}_overlap_{lvl}.xlsx"
        _save_formatted_overlap_table(level_df, lvl, formatted_path)

    return df


# =============================================================================
# Excel export helper
# =============================================================================

def _save_formatted_overlap_table(level_df, level, output_path):
    """
    Export a formatted, human-readable embedding overlap analysis table to Excel.

    Generates a styled Excel workbook with:
      - A merged title row identifying the variation level
      - A styled header row with white text on dark blue background
      - Alternating row fills (white / light blue) for readability
      - A bold summary row with mean ± std for all three metrics
      - Fixed column widths optimized for the typical command string lengths

    This is an internal helper function called once per variation level
    by `analyze_embeddings`. It is prefixed with `_` to indicate it is
    not part of the public API.

    Parameters
    ----------
    level_df : pandas.DataFrame
        Filtered DataFrame containing only rows for the given variation level.
        Must have columns: 'default_command', 'variation_command',
        'cosine_similarity', 'euclidean_distance', 'levenshtein_distance'.
        Must be zero-indexed (call `.reset_index(drop=True)` before passing).
    level : str
        Variation level identifier ("l1", "l2", or "l3").
        Used for the worksheet title and the Excel title row.
    output_path : str
        Full path (including filename) where the `.xlsx` file will be saved.

    Returns
    -------
    None
        Saves the workbook to `output_path` and prints a confirmation message.

    Notes
    -----
    Row Layout:
      - Row 1 : Merged title cell spanning all 6 columns
      - Row 2 : Column headers
      - Rows 3 to (N+2) : One data row per task (N = len(level_df))
      - Row (N+3) : AVERAGE summary row with mean ± std strings

    Number Format:
      Metric columns (4–6) use the "0.0000" Excel number format for
      consistent 4-decimal-place display. However, the AVERAGE row uses
      pre-formatted "XX.XXXX ± YY.YYYY" strings rather than numeric values,
      so they are stored as text cells.

    Alternating Row Fills:
      Odd-indexed rows (i % 2 == 1, 0-based) receive a light blue fill
      (`fgColor="DCE6F1"`). Even-indexed rows have no fill (None), resulting
      in the default white background.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter  # Converts column index to letter (1 → "A")

    wb = openpyxl.Workbook()  # Create a new empty workbook
    ws = wb.active            # Get the default active worksheet
    ws.title = f"{level.upper()} Variations"  # Name the worksheet tab

    # ── Row 1: Merged title cell ───────────────────────────────────────────
    title = f"Overlapping Analysis - {level.upper()} Variations"
    ws.merge_cells("A1:F1")   # Merge columns A through F into a single cell
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = Font(bold=True, size=13)            # Large bold font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22  # Slightly taller row for visual prominence

    # ── Row 2: Column headers ─────────────────────────────────────────────
    headers = [
        "N°",                      # 1-based row index
        "Original Task Command",   # Default command string
        "Variation Task Command",  # L1/L2/L3 variation command string
        "Cosine Similarity",        # Semantic: angular distance metric
        "Euclidean Dist.",          # Semantic: L2 distance metric
        "Levenshtein Dist."         # Lexical: normalized edit distance
    ]

    # Header styling constants
    header_fill = PatternFill(fill_type="solid", fgColor="2F5496")  # Dark blue background
    header_font = Font(bold=True, color="FFFFFF")                    # White bold text
    thin = Side(style="thin", color="000000")                        # Thin black border side
    border = Border(left=thin, right=thin, top=thin, bottom=thin)    # All-sides thin border

    for col, h in enumerate(headers, start=1):  # 1-based column indexing for openpyxl
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    ws.row_dimensions[2].height = 20  # Fixed header row height

    # ── Rows 3 to N+2: Data rows ───────────────────────────────────────────
    # Light blue fill for alternating rows (applied to odd-indexed rows, 0-based)
    alt_fill = PatternFill(fill_type="solid", fgColor="DCE6F1")
    num_fmt_4 = "0.0000"  # Excel number format: always show 4 decimal places

    for i, row in level_df.iterrows():
        excel_row = i + 3  # Shift by 3: row 1 = title, row 2 = headers, row 3+ = data

        # Apply alternating row fill: odd i → light blue, even i → no fill (white)
        fill = alt_fill if i % 2 == 1 else None

        # Column values for this data row
        values = [
            i + 1,                                      # Column 1: 1-based row number
            row['default_command'],                      # Column 2: original task command
            row['variation_command'],                    # Column 3: variation command
            round(row['cosine_similarity'], 4),          # Column 4: cosine sim (4 decimals)
            round(row['euclidean_distance'], 4),         # Column 5: euclidean dist (4 decimals)
            round(row['levenshtein_distance'], 4),       # Column 6: levenshtein dist (4 decimals)
        ]

        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            if fill:
                cell.fill = fill  # Apply alternating row color

            # Apply 4-decimal number format to metric columns (4, 5, 6)
            if col >= 4:
                cell.number_format = num_fmt_4
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Final row: AVERAGE summary ─────────────────────────────────────────
    avg_row = len(level_df) + 3  # Row index immediately after the last data row
    avg_fill = PatternFill(fill_type="solid", fgColor="F4B942")  # Golden/amber background
    avg_font = Font(bold=True)

    def fmt_mean_std(col_name):
        """
        Compute and format 'mean ± std' string for a DataFrame column.

        Parameters
        ----------
        col_name : str
            Column name in `level_df` to summarize.

        Returns
        -------
        str
            Formatted string "X.XXXX ± Y.YYYY" representing the
            arithmetic mean and sample standard deviation of the column.
        """
        mean = level_df[col_name].mean()   # Arithmetic mean across all tasks
        std = level_df[col_name].std()     # Sample std (pandas default: ddof=1)
        return f"{mean:.4f} ± {std:.4f}"  # Format to 4 decimal places

    # Build the AVERAGE row: label in col 1, blanks for text cols, stats for metric cols
    avg_values = [
        "AVERAGE", "", "",                            # Col 1-3: label + empty text cells
        fmt_mean_std('cosine_similarity'),             # Col 4: mean ± std of cosine sim
        fmt_mean_std('euclidean_distance'),            # Col 5: mean ± std of euclidean dist
        fmt_mean_std('levenshtein_distance'),          # Col 6: mean ± std of levenshtein dist
    ]

    for col, val in enumerate(avg_values, start=1):
        cell = ws.cell(row=avg_row, column=col, value=val)
        cell.font = avg_font
        cell.fill = avg_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Column widths (manually tuned for typical command string lengths) ──
    col_widths = [6, 52, 52, 18, 16, 18]  # [N°, OrigCmd, VarCmd, CosSim, EucDist, LevDist]
    for col, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # ── Save workbook ──────────────────────────────────────────────────────
    wb.save(output_path)
    print(f"✓ Formatted overlap table saved: {output_path}")


# =============================================================================
# Per-rollout diagnostic analysis
# =============================================================================

def compare_per_rollout_similarity(embeddings, task_id=0, level='l1'):
    """
    Compute and report cosine similarity and Euclidean distance between default
    and variation embeddings on a per-rollout basis for a specific task.

    While `analyze_embeddings` uses a single mean embedding per command (averaged
    over all rollouts and steps), this function operates at the rollout level:
    it compares the default and variation embedding for each individual rollout
    episode. This allows assessment of:
      - **Intra-condition variability**: How consistent are embeddings within
        the same command across rollouts? (std over rollouts)
      - **Cross-condition alignment**: Does the default-variation similarity
        hold consistently across episodes, or does it vary with episode outcome?

    Parameters
    ----------
    embeddings : dict
        Dictionary of embedding records (output of `load_embeddings()`).
        Must contain keys "task_{task_id:02d}_default" and
        "task_{task_id:02d}_{level}". Each record must have the
        "embedding_per_rollout" key with shape (N_rollouts, embed_dim).
    task_id : int, optional
        0-based task index to analyze. Default: 0.
    level : str, optional
        Variation level to compare against default ("l1", "l2", or "l3").
        Default: "l1".

    Returns
    -------
    tuple of (numpy.ndarray, numpy.ndarray) or None
        A tuple (similarities, euc_dists) where:
          - similarities : shape (min_rollouts,) float array of per-rollout
                           cosine similarities between default and variation.
          - euc_dists    : shape (min_rollouts,) float array of per-rollout
                           Euclidean (L2) distances between default and variation.
        Returns None if the required keys are missing or per-rollout embeddings
        are unavailable.

    Notes
    -----
    Rollout Pairing:
      The i-th default rollout is compared against the i-th variation rollout.
      If the number of rollouts differs between default and variation (e.g., due
      to partial data), only `min(N_default, N_variation)` pairs are compared.
      The minimum count is reported in the console output.

    Interpretation:
      - High mean similarity with low std suggests the model encodes the two
        commands consistently across all episodes (robust language grounding).
      - Low mean similarity or high std suggests the model's representation is
        command-sensitive or episode-outcome-sensitive.
    """
    # Construct the expected dictionary keys from task_id and level
    default_key = f"task_{task_id:02d}_default"  # e.g., "task_03_default"
    var_key = f"task_{task_id:02d}_{level}"      # e.g., "task_03_l2"

    # Guard: both keys must exist in the embeddings dictionary
    if default_key not in embeddings or var_key not in embeddings:
        print(f"Keys not found: {default_key} or {var_key}")
        return None

    default_data = embeddings[default_key]  # Full embedding record for default
    var_data = embeddings[var_key]          # Full embedding record for variation

    # Guard: per-rollout embeddings must be available in both records
    if 'embedding_per_rollout' not in default_data or 'embedding_per_rollout' not in var_data:
        print("Per-rollout embeddings not available")
        return None

    # Extract per-rollout embedding arrays: shape (N_rollouts, embed_dim)
    default_rollouts = default_data['embedding_per_rollout']
    var_rollouts = var_data['embedding_per_rollout']

    # Use the smaller rollout count to avoid index-out-of-bounds comparisons
    min_rollouts = min(len(default_rollouts), len(var_rollouts))

    # ── Compute per-rollout cosine similarities ────────────────────────────
    similarities = []
    for i in range(min_rollouts):
        # Compare default rollout i vs. variation rollout i
        # sklearn expects 2D inputs: wrap each 1D vector in a list
        sim = cosine_similarity([default_rollouts[i]], [var_rollouts[i]])[0, 0]
        similarities.append(sim)

    similarities = np.array(similarities)  # Convert list to numpy array for statistics

    # ── Compute per-rollout Euclidean distances ────────────────────────────
    euc_dists = []
    for i in range(min_rollouts):
        dist = euclidean_distances([default_rollouts[i]], [var_rollouts[i]])[0, 0]
        euc_dists.append(dist)

    euc_dists = np.array(euc_dists)  # Convert list to numpy array for statistics

    # ── Console report ────────────────────────────────────────────────────
    print(f"\nTask {task_id} - Default vs {level.upper()}")
    print(f"  Default command:   {default_data['command_text']}")
    print(f"  Variation command: {var_data['command_text']}")

    # Per-rollout cosine similarity statistics
    print(f"\n  Per-rollout cosine similarity ({min_rollouts} rollouts):")
    print(f"    Mean:   {similarities.mean():.4f}")  # Average alignment across rollouts
    print(f"    Std:    {similarities.std():.4f}")   # Consistency across rollouts
    print(f"    Min:    {similarities.min():.4f}")   # Worst-case alignment
    print(f"    Max:    {similarities.max():.4f}")   # Best-case alignment

    # Per-rollout Euclidean distance statistics
    print(f"\n  Per-rollout euclidean distance L2 ({min_rollouts} rollouts):")
    print(f"    Mean:   {euc_dists.mean():.4f}")
    print(f"    Std:    {euc_dists.std():.4f}")
    print(f"    Min:    {euc_dists.min():.4f}")
    print(f"    Max:    {euc_dists.max():.4f}")

    return similarities, euc_dists


# =============================================================================
# CLI entry point
# =============================================================================

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description=(
            "Analyze TinyVLA command variations with rollout-based "
            "pre-action-head embeddings"
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
        # epilog: displayed verbatim at the end of --help output
        epilog="""
Examples:
  # Single file
  python analyze_embedding_overlap_rollout.py --embedding_file embeddings.pkl

  # Multiple files (one per command level)
  python analyze_embedding_overlap_rollout.py --embedding_files \\
      .../rollout_embeddings_tinyvla_libero_goal_default_first_step_r10.pkl \\
      .../rollout_embeddings_tinyvla_libero_goal_l1_first_step_r10.pkl \\
      .../rollout_embeddings_tinyvla_libero_goal_l2_first_step_r10.pkl \\
      .../rollout_embeddings_tinyvla_libero_goal_l3_first_step_r10.pkl

  # From directory
  python analyze_embedding_overlap_rollout.py \\
      --embedding_dir /mnt/beegfs/a.cardamone7/outputs/embeddings/tinyvla
"""
    )

    # ── Input source arguments ─────────────────────────────────────────────
    parser.add_argument(
        "--embedding_file",
        type=str,
        default=None,
        help="Path to a single embeddings pickle file containing all levels"
    )
    parser.add_argument(
        "--embedding_files",
        type=str,
        nargs="+",      # Accepts one or more file paths
        default=None,
        help="Paths to multiple pickle files (one per command level)"
    )
    parser.add_argument(
        "--embedding_dir",
        type=str,
        default=None,
        help="Directory containing embedding pickle files (auto-discovers all *.pkl)"
    )

    # ── Output argument ────────────────────────────────────────────────────
    parser.add_argument(
        "--output_csv",
        type=str,
        default=None,
        help="Output CSV path (default: auto-generated from input path)"
    )

    # ── Per-rollout analysis arguments ─────────────────────────────────────
    parser.add_argument(
        "--per_rollout_analysis",
        action="store_true",  # Boolean flag: True if present, False if absent
        help="Run per-rollout similarity analysis for a specific task"
    )
    parser.add_argument(
        "--task_id",
        type=int,
        default=0,
        help="Task ID (0-based) for per-rollout analysis (default: 0)"
    )
    parser.add_argument(
        "--level",
        type=str,
        default="l1",
        help="Variation level for per-rollout analysis: l1, l2, or l3 (default: l1)"
    )

    args = parser.parse_args()

    # ── Determine the embedding source and load data ───────────────────────
    if args.embedding_file:
        # Single-file mode: load the pickle directly
        with open(args.embedding_file, 'rb') as f:
            embeddings = pickle.load(f)
        print(f"Loaded {len(embeddings)} embeddings from {args.embedding_file}")
        # Use the input file path as the base for output naming
        output_base = args.embedding_file

    elif args.embedding_files or args.embedding_dir:
        # Multi-file or directory mode: delegate to load_embeddings()
        embeddings = load_embeddings(
            embedding_files=args.embedding_files,
            embedding_dir=args.embedding_dir
        )
        # Set output base: directory path + "combined_analysis" stem
        output_base = args.embedding_dir or os.path.dirname(args.embedding_files[0])
        output_base = os.path.join(output_base, "combined_analysis")

    else:
        # Fallback default: use the TinyVLA embeddings directory on the cluster
        default_dir = "/mnt/beegfs/a.cardamone7/outputs/embeddings/tinyvla/"
        embeddings = load_embeddings(embedding_dir=default_dir)
        output_base = os.path.join(default_dir, "combined_analysis")

    # ── Run the main analysis ──────────────────────────────────────────────
    # Use --output_csv if provided, otherwise auto-generate from output_base
    df = analyze_embeddings(
        embeddings,
        output_csv=args.output_csv or f"{output_base}.csv"
    )

    # ── Optionally run per-rollout diagnostic analysis ─────────────────────
    if args.per_rollout_analysis:
        similarities = compare_per_rollout_similarity(
            embeddings,
            task_id=args.task_id,
            level=args.level
        )