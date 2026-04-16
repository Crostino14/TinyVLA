#!/usr/bin/env python3
"""
Build TinyVLA evaluation tables from a directory of .txt logs.

What it does:
- scans one input folder recursively
- auto-separates L1 and L2 files
- parses batch logs and single-task reruns
- merges results by (level, task, variant, seed)
- single-task reruns override batch logs
- newer timestamps override older ones
- writes:
    * tinyvla_l1_evaluation.xlsx
    * tinyvla_l2_evaluation.xlsx
    * tinyvla_evaluation_tables.xlsx   (L1 + L2 + Info)

Usage:
    python build_tinyvla_tables.py --input-dir /path/to/logs
    python build_tinyvla_tables.py --input-dir /path/to/logs --output-dir /path/to/out --debug
"""

import re
import math
import argparse
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


TASK_MAPPING = {
    "Open the middle layer of the drawer": 1,
    "Draw out the middle drawer of the cabinet": 1,
    "Slide out the middle drawer of the cabinet": 1,
    "Ease out the middle drawer of the cabinet": 1,

    "Put the bowl on the stove": 2,
    "Position the bowl on the stove": 2,
    "Rest the bowl on the stove": 2,
    "Lay the bowl on the stove": 2,

    "Put the wine bottle on the top of the cabinet": 3,
    "Position the wine bottle on the top of the cabinet": 3,
    "Rest the wine bottle on the top of the cabinet": 3,
    "Lay the wine bottle on the top of the cabinet": 3,

    "Open the top layer of the drawer and put the bowl inside": 4,
    "Draw out the top drawer of the cabinet and position the bowl inside": 4,
    "Slide out the top drawer of the cabinet and rest the bowl inside": 4,
    "Ease out the top drawer of the cabinet and lay the bowl inside": 4,

    "Put the bowl on the top of the cabinet": 5,
    "Position the bowl on the top of the cabinet": 5,
    "Rest the bowl on the top of the cabinet": 5,
    "Lay the bowl on the top of the cabinet": 5,

    "Push the plate to the front of the stove": 6,
    "Slide the plate to the front of the stove": 6,
    "Shift the plate to the front of the stove": 6,
    "Guide the plate to the front of the stove": 6,

    "Put the cream cheese on the bowl": 7,
    "Position the cream cheese on the bowl": 7,
    "Rest the cream cheese on the bowl": 7,
    "Lay the cream cheese on the bowl": 7,

    "Turn on the stove": 8,
    "Ignite the stove": 8,
    "Start the stove": 8,
    "Power on the stove": 8,

    "Put the bowl on the plate": 9,
    "Position the bowl on the plate": 9,
    "Rest the bowl on the plate": 9,
    "Lay the bowl on the plate": 9,

    "Put the wine bottle on the rack": 10,
    "Position the wine bottle on the rack": 10,
    "Rest the wine bottle on the rack": 10,
    "Lay the wine bottle on the rack": 10,

    "The middle drawer of the cabinet needs to be opened": 1,
    "The middle drawer of the cabinet should be opened": 1,
    "Let the middle drawer of the cabinet be opened": 1,
    "The middle drawer layer must be opened": 1,

    "The stove needs to have the bowl on it": 2,
    "The bowl should be put on the stove": 2,
    "Let the bowl be put on the stove": 2,
    "The bowl must be put on the stove": 2,

    "The top of the cabinet needs to have the wine bottle on it": 3,
    "The wine bottle should be put on the top of the cabinet": 3,
    "Let the wine bottle be put on top of the cabinet": 3,
    "The wine bottle must be put on the cabinet's top": 3,

    "The top drawer of the cabinet needs to be opened and the bowl needs to be put inside": 4,
    "The top drawer of the cabinet should be opened and the bowl should be put inside": 4,
    "Let the top drawer of the cabinet be opened and the bowl be put inside": 4,
    "The top drawer of the cabinet must be opened and the bowl must be put inside": 4,

    "The top of the cabinet needs to have the bowl on it": 5,
    "The bowl should be put on the top of the cabinet": 5,
    "Let the bowl be put on top of the cabinet": 5,
    "The bowl must be put on the cabinet's top": 5,

    "The space in front of the stove needs to have the plate in it": 6,
    "The plate should be pushed to the front of the stove": 6,
    "Let the plate be pushed to the front of the stove": 6,
    "The plate must be pushed to the stove's front": 6,

    "The cream cheese needs to be put on the bowl": 7,
    "The cream cheese should be put on the bowl": 7,
    "Let the cream cheese be put on the bowl": 7,
    "The cream cheese must be put on the bowl": 7,

    "The stove needs to be turned on": 8,
    "The stove should be turned on": 8,
    "Let the stove be turned on": 8,
    "The stove must be turned on": 8,

    "The plate needs to have the bowl on it": 9,
    "The bowl should be put on the plate": 9,
    "Let the bowl be put on the plate": 9,
    "The bowl must be put on the plate": 9,

    "The rack needs to be filled with the wine bottle in it": 10,
    "The wine bottle should be put on the rack": 10,
    "Let the wine bottle be put on the rack": 10,
    "The wine bottle must be put on the rack": 10,
}

VARIANT_BY_COMMAND = {
    1: {
        "Open the middle layer of the drawer": "Original",
        "Draw out the middle drawer of the cabinet": "V1",
        "Slide out the middle drawer of the cabinet": "V2",
        "Ease out the middle drawer of the cabinet": "V3",
        "The middle drawer of the cabinet needs to be opened": "Original",
        "The middle drawer of the cabinet should be opened": "V1",
        "Let the middle drawer of the cabinet be opened": "V2",
        "The middle drawer layer must be opened": "V3",
    },
    2: {
        "Put the bowl on the stove": "Original",
        "Position the bowl on the stove": "V1",
        "Rest the bowl on the stove": "V2",
        "Lay the bowl on the stove": "V3",
        "The stove needs to have the bowl on it": "Original",
        "The bowl should be put on the stove": "V1",
        "Let the bowl be put on the stove": "V2",
        "The bowl must be put on the stove": "V3",
    },
    3: {
        "Put the wine bottle on the top of the cabinet": "Original",
        "Position the wine bottle on the top of the cabinet": "V1",
        "Rest the wine bottle on the top of the cabinet": "V2",
        "Lay the wine bottle on the top of the cabinet": "V3",
        "The top of the cabinet needs to have the wine bottle on it": "Original",
        "The wine bottle should be put on the top of the cabinet": "V1",
        "Let the wine bottle be put on top of the cabinet": "V2",
        "The wine bottle must be put on the cabinet's top": "V3",
    },
    4: {
        "Open the top layer of the drawer and put the bowl inside": "Original",
        "Draw out the top drawer of the cabinet and position the bowl inside": "V1",
        "Slide out the top drawer of the cabinet and rest the bowl inside": "V2",
        "Ease out the top drawer of the cabinet and lay the bowl inside": "V3",
        "The top drawer of the cabinet needs to be opened and the bowl needs to be put inside": "Original",
        "The top drawer of the cabinet should be opened and the bowl should be put inside": "V1",
        "Let the top drawer of the cabinet be opened and the bowl be put inside": "V2",
        "The top drawer of the cabinet must be opened and the bowl must be put inside": "V3",
    },
    5: {
        "Put the bowl on the top of the cabinet": "Original",
        "Position the bowl on the top of the cabinet": "V1",
        "Rest the bowl on the top of the cabinet": "V2",
        "Lay the bowl on the top of the cabinet": "V3",
        "The top of the cabinet needs to have the bowl on it": "Original",
        "The bowl should be put on the top of the cabinet": "V1",
        "Let the bowl be put on top of the cabinet": "V2",
        "The bowl must be put on the cabinet's top": "V3",
    },
    6: {
        "Push the plate to the front of the stove": "Original",
        "Slide the plate to the front of the stove": "V1",
        "Shift the plate to the front of the stove": "V2",
        "Guide the plate to the front of the stove": "V3",
        "The space in front of the stove needs to have the plate in it": "Original",
        "The plate should be pushed to the front of the stove": "V1",
        "Let the plate be pushed to the front of the stove": "V2",
        "The plate must be pushed to the stove's front": "V3",
    },
    7: {
        "Put the cream cheese on the bowl": "Original",
        "Position the cream cheese on the bowl": "V1",
        "Rest the cream cheese on the bowl": "V2",
        "Lay the cream cheese on the bowl": "V3",
        "The cream cheese needs to be put on the bowl": "Original",
        "The cream cheese should be put on the bowl": "V1",
        "Let the cream cheese be put on the bowl": "V2",
        "The cream cheese must be put on the bowl": "V3",
    },
    8: {
        "Turn on the stove": "Original",
        "Ignite the stove": "V1",
        "Start the stove": "V2",
        "Power on the stove": "V3",
        "The stove needs to be turned on": "Original",
        "The stove should be turned on": "V1",
        "Let the stove be turned on": "V2",
        "The stove must be turned on": "V3",
    },
    9: {
        "Put the bowl on the plate": "Original",
        "Position the bowl on the plate": "V1",
        "Rest the bowl on the plate": "V2",
        "Lay the bowl on the plate": "V3",
        "The plate needs to have the bowl on it": "Original",
        "The bowl should be put on the plate": "V1",
        "Let the bowl be put on the plate": "V2",
        "The bowl must be put on the plate": "V3",
    },
    10: {
        "Put the wine bottle on the rack": "Original",
        "Position the wine bottle on the rack": "V1",
        "Rest the wine bottle on the rack": "V2",
        "Lay the wine bottle on the rack": "V3",
        "The rack needs to be filled with the wine bottle in it": "Original",
        "The wine bottle should be put on the rack": "V1",
        "Let the wine bottle be put on the rack": "V2",
        "The wine bottle must be put on the rack": "V3",
    },
}

TINYVLA_L1_ORIGINAL_VALUES = {
    1: [94.7, 94.7, 94.7],
    2: [89.3, 89.3, 89.3],
    3: [100.0, 100.0, 100.0],
    4: [86.7, 86.7, 86.7],
    5: [92.7, 92.7, 92.7],
    6: [74.7, 74.7, 74.7],
    7: [76.0, 76.0, 76.0],
    8: [97.3, 97.3, 97.3],
    9: [89.3, 89.3, 89.3],
    10: [83.3, 83.3, 83.3],
}

TINYVLA_L2_ORIGINAL_VALUES = {
    1: [95.3, 95.3, 95.3],
    2: [60.0, 60.0, 60.0],
    3: [2.7, 2.7, 2.7],
    4: [76.0, 76.0, 76.0],
    5: [0.0, 0.0, 0.0],
    6: [11.3, 11.3, 11.3],
    7: [90.0, 90.0, 90.0],
    8: [100.0, 100.0, 100.0],
    9: [20.7, 20.7, 20.7],
    10: [48.0, 48.0, 48.0],
}


def timestamp_from_filename(name: str) -> str:
    m = re.search(r'(\d{4}_\d{2}_\d{2}-\d{2}_\d{2}_\d{2})', name)
    return m.group(1) if m else "0000_00_00-00_00_00"


def extract_seed_from_filename(path: str):
    m = re.search(r'_seed(\d+)', Path(path).name)
    return int(m.group(1)) if m else None


def detect_file_scope(path: str):
    name = Path(path).name

    m_range = re.search(r'_tasks(\d+)-(\d+)', name)
    if m_range:
        start0 = int(m_range.group(1))
        end0 = int(m_range.group(2))
        return "batch", list(range(start0 + 1, end0 + 2))

    m_all = re.findall(r'_task(\d+)', name)
    if m_all:
        task0 = int(m_all[-1])
        return "single", [task0 + 1]

    return "unknown", []


def infer_level_from_filename(path: str):
    name = Path(path).name.lower()
    if "--l1" in name or "_l1_" in name or "l1_variations" in name:
        return "L1"
    if "--l2" in name or "_l2_" in name or "l2_variations" in name:
        return "L2"
    return None


def compute_stats(values):
    if not values:
        return None, None
    mean = sum(values) / len(values)
    if len(values) == 1:
        return mean, 0.0
    var = sum((x - mean) ** 2 for x in values) / (len(values) - 1)
    return mean, math.sqrt(var)


def format_value(mean, std):
    if mean is None:
        return "N/A"
    return f"{mean:.1f}% ± {std:.1f}%".replace(".", ",")


def parse_txt_evaluation(filepath, level=None):
    filename = Path(filepath).name
    seed = extract_seed_from_filename(filepath)
    file_scope_type, file_scope_tasks = detect_file_scope(filepath)
    timestamp = timestamp_from_filename(filename)

    with open(filepath, "r", encoding="utf-8") as f:
        content = f.read()

    if level:
        sec_pat = re.compile(r'={80}\s*\nEVALUATING:\s*(L[12])\s*\n={80}', re.M)
        matches = list(sec_pat.finditer(content))
        kept = []
        for i, m in enumerate(matches):
            sec_level = m.group(1)
            start = m.end()
            end = matches[i + 1].start() if i + 1 < len(matches) else len(content)
            if sec_level == level:
                kept.append(content[start:end])
        if kept:
            content = "\n".join(kept)

    blocks = re.split(r'(?=Testing VERSION:)', content)
    parsed = []

    for block in blocks:
        if "Testing VERSION:" not in block:
            continue

        version_match = re.search(r'Testing VERSION:\s*(v?\d+)', block)
        orig_match = re.search(r'Original Command:\s*(.+?)\n', block)
        var_match = re.search(r'Variation Command:\s*(.+?)\n', block)

        if not orig_match or not var_match:
            continue

        version_label = version_match.group(1) if version_match else None
        original_cmd = orig_match.group(1).strip()
        variation_cmd = var_match.group(1).strip()

        rate = None
        final_rate_match = re.search(r'Success Rate:\s*([\d.]+)%', block)
        if final_rate_match:
            rate = float(final_rate_match.group(1))
        else:
            episodic = re.findall(r'Total successes:\s*\d+\s*\(([\d.]+)%\)', block)
            if episodic:
                rate = float(episodic[-1])

        if rate is None:
            continue

        task_num = TASK_MAPPING.get(variation_cmd) or TASK_MAPPING.get(original_cmd)
        if task_num is None and len(file_scope_tasks) == 1:
            task_num = file_scope_tasks[0]
        if task_num is None:
            continue

        variant = VARIANT_BY_COMMAND.get(task_num, {}).get(variation_cmd)
        if variant is None and version_label:
            vm = re.search(r'(\d+)$', version_label)
            if vm:
                idx = int(vm.group(1))
                variant = {1: "V1", 2: "V2", 3: "V3"}.get(idx)
        if variant is None:
            continue

        parsed.append({
            "level": level,
            "task": task_num,
            "variant": variant,
            "seed": seed,
            "rate": rate,
            "filename": filename,
            "timestamp": timestamp,
            "source_kind": file_scope_type,
            "original_cmd": original_cmd,
            "variation_cmd": variation_cmd,
        })

    return parsed


def better_candidate(old, new):
    if old is None:
        return True

    old_single = old["source_kind"] == "single"
    new_single = new["source_kind"] == "single"

    if new_single and not old_single:
        return True
    if old_single and not new_single:
        return False

    return new["timestamp"] > old["timestamp"]


def aggregate_runs(parsed_runs):
    selected = {}
    for run in parsed_runs:
        key = (run["task"], run["variant"], run["seed"])
        if better_candidate(selected.get(key), run):
            selected[key] = run

    aggregated = defaultdict(lambda: defaultdict(dict))
    for (task, variant, seed), run in selected.items():
        aggregated[task][variant][seed] = run["rate"]

    return selected, aggregated


def style_sheet(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 18
    for col in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 22
    ws.row_dimensions[1].height = 35


def fill_level_sheet(ws, level, aggregated_data):
    original_rates = TINYVLA_L1_ORIGINAL_VALUES if level == "L1" else TINYVLA_L2_ORIGINAL_VALUES
    variant_names = ["Original", "V1", "V2", "V3"]

    header = ["Task"] + [f"{v}\nMean SR% ± std%" for v in variant_names] + ["Mean of means\nSR% ± std%"]
    ws.append(header)

    overall_variant_task_means = {v: [] for v in variant_names}
    issues = []

    for task_num in range(1, 11):
        row = [f"Task {task_num}"]
        per_variant_means_this_task = []

        for variant in variant_names:
            if variant == "Original":
                rates = original_rates.get(task_num)
                if rates:
                    mean, std = compute_stats(rates)
                    row.append(format_value(mean, std))
                    overall_variant_task_means[variant].append(mean)
                    per_variant_means_this_task.append(mean)
                else:
                    row.append("N/A")
                continue

            seed_map = aggregated_data.get(task_num, {}).get(variant, {})
            if seed_map:
                ordered_rates = [seed_map[s] for s in sorted(seed_map)]
                mean, std = compute_stats(ordered_rates)

                missing = [s for s in [0, 1, 2] if s not in seed_map]
                cell_value = format_value(mean, std)
                if missing:
                    cell_value += f" (!{missing})"
                    issues.append(f"{level} Task {task_num} {variant}: missing seeds {missing}")

                row.append(cell_value)
                overall_variant_task_means[variant].append(mean)
                per_variant_means_this_task.append(mean)
            else:
                row.append("N/A")

        task_mean, task_std = compute_stats(per_variant_means_this_task)
        row.append(format_value(task_mean, task_std))
        ws.append(row)

    final_row = ["OVERALL"]
    all_variant_task_means = []
    for variant in variant_names:
        vals = overall_variant_task_means[variant]
        mean, std = compute_stats(vals)
        final_row.append(format_value(mean, std))
        all_variant_task_means.extend(vals)

    overall_mean, overall_std = compute_stats(all_variant_task_means)
    final_row.append(format_value(overall_mean, overall_std))
    ws.append(final_row)

    style_sheet(ws)
    return issues


def create_single_level_workbook(level, aggregated_data, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = level
    issues = fill_level_sheet(ws, level, aggregated_data)
    wb.save(output_path)
    return issues


def create_combined_workbook(l1_agg, l2_agg, info_lines, output_path):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "L1"
    fill_level_sheet(ws1, "L1", l1_agg)

    ws2 = wb.create_sheet("L2")
    fill_level_sheet(ws2, "L2", l2_agg)

    ws3 = wb.create_sheet("Info")
    ws3.append(["TinyVLA aggregation info"])
    ws3["A1"].font = Font(bold=True)
    ws3.column_dimensions["A"].width = 120
    for line in info_lines:
        ws3.append([line])

    wb.save(output_path)


def collect_txt_files(input_dir: Path):
    return sorted([p for p in input_dir.rglob("*.txt") if p.is_file()])


def split_files_by_level(files):
    l1_files, l2_files, unknown = [], [], []
    for f in files:
        level = infer_level_from_filename(str(f))
        if level == "L1":
            l1_files.append(str(f))
        elif level == "L2":
            l2_files.append(str(f))
        else:
            unknown.append(str(f))
    return l1_files, l2_files, unknown


def parse_many(files, level):
    runs = []
    for f in files:
        runs.extend(parse_txt_evaluation(f, level=level))
    return runs


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-dir", required=True, help="Folder containing TinyVLA .txt logs")
    parser.add_argument("--output-dir", default=".", help="Destination folder for .xlsx files")
    parser.add_argument("--prefix", default="tinyvla", help="Output filename prefix")
    parser.add_argument("--debug", action="store_true")
    args = parser.parse_args()

    input_dir = Path(args.input_dir).expanduser().resolve()
    output_dir = Path(args.output_dir).expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    if not input_dir.exists() or not input_dir.is_dir():
        raise SystemExit(f"Input directory not found: {input_dir}")

    all_txt = collect_txt_files(input_dir)
    l1_files, l2_files, unknown = split_files_by_level(all_txt)

    print("=" * 72)
    print("BUILD TINYVLA EVALUATION TABLES")
    print("=" * 72)
    print(f"Input dir : {input_dir}")
    print(f"Output dir: {output_dir}")
    print(f"Found txt : {len(all_txt)}")
    print(f"L1 files  : {len(l1_files)}")
    print(f"L2 files  : {len(l2_files)}")
    print(f"Unknown   : {len(unknown)}")

    if args.debug and unknown:
        print("\n[DEBUG] Files with unknown level:")
        for f in unknown:
            print("  ", f)

    l1_runs = parse_many(l1_files, "L1")
    l2_runs = parse_many(l2_files, "L2")

    l1_selected, l1_agg = aggregate_runs(l1_runs)
    l2_selected, l2_agg = aggregate_runs(l2_runs)

    if args.debug:
        print(f"\n[DEBUG] Parsed L1 runs: {len(l1_runs)} | selected unique keys: {len(l1_selected)}")
        for k in sorted(l1_selected):
            r = l1_selected[k]
            print("  L1", k, r["rate"], r["filename"])
        print(f"\n[DEBUG] Parsed L2 runs: {len(l2_runs)} | selected unique keys: {len(l2_selected)}")
        for k in sorted(l2_selected):
            r = l2_selected[k]
            print("  L2", k, r["rate"], r["filename"])

    l1_path = output_dir / f"{args.prefix}_l1_evaluation.xlsx"
    l2_path = output_dir / f"{args.prefix}_l2_evaluation.xlsx"
    combined_path = output_dir / f"{args.prefix}_evaluation_tables.xlsx"

    l1_issues = create_single_level_workbook("L1", l1_agg, l1_path)
    l2_issues = create_single_level_workbook("L2", l2_agg, l2_path)

    info_lines = [
        f"Input directory: {input_dir}",
        f"Total .txt files found: {len(all_txt)}",
        f"L1 files: {len(l1_files)}",
        f"L2 files: {len(l2_files)}",
        "Merge policy:",
        "- key = (level, task, variant, seed)",
        "- single-task reruns override batch files",
        "- if two runs have same scope, newest timestamp wins",
        "- final task column is the mean of the available variant means (Original, V1, V2, V3)",
    ]
    if unknown:
        info_lines.append(f"Unknown-level files ignored: {len(unknown)}")
        info_lines.extend([f"- {Path(x).name}" for x in unknown[:20]])
    if l1_issues or l2_issues:
        info_lines.append("Missing seed warnings:")
        info_lines.extend([f"- {x}" for x in (l1_issues + l2_issues)])

    create_combined_workbook(l1_agg, l2_agg, info_lines, combined_path)

    print(f"\n[OK] {l1_path}")
    print(f"[OK] {l2_path}")
    print(f"[OK] {combined_path}")

    if l1_issues or l2_issues:
        print("\nWarnings:")
        for x in l1_issues + l2_issues:
            print(" -", x)


if __name__ == "__main__":
    main()
